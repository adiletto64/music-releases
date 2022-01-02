import openpyxl
from .models import Release
from django_countries.data import COUNTRIES

from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError
from releases.models.release import BASE_STYLE_CHOICES


FORMATS = Release.Formats.values
STYLES = BASE_STYLE_CHOICES

COUNTRIES_DICT = {full_name: short_name for short_name, full_name in COUNTRIES.items()}


def save_excel_file(file, profile):
    """
    Parse, validate and save file. If there is error returns it, otherwise returns None
    """

    # you may say that file can be not excel format,
    # but i added file validation in ReleaseImportForm (forms.py)
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):

        # as django_countries saves country in DB with 2 chars ("New Zealand" -> "NZ")
        # and countries dict looks like {"NZ": "New Zealand"}
        valid_country = COUNTRIES_DICT[sheet.cell(row, 3).value]
        if not valid_country:
            return f"Wrong country at {row} row, 3 column"

        try:
            # Convert DD.MM.YYYY format to YYYY-MM-DD
            valid_date = sheet.cell(row, 4).value.strftime("%Y-%m-%d")
        except AttributeError:
            return f"wrong date format at {row} row, 4 column"

        try:
            label_name = sheet.cell(row, 9).value
            label = profile.labels.get(name=label_name)
        except ObjectDoesNotExist:
            return f"You haven't label named {label_name}. Error at {row} row, 9 column"

        format = sheet.cell(row, 6).value
        style = sheet.cell(row, 8).value

        try:
            release = Release.objects.create(
                band_name=sheet.cell(row, 1).value,
                album_title=sheet.cell(row, 2).value,
                country=valid_country,
                release_date=valid_date,
                media_format_details=sheet.cell(row, 5).value,
                format=format if (format in FORMATS) else "Other",
                limited_edition=sheet.cell(row, 7).value,
                base_style=style if (style in STYLES) else "Other",
                profile=profile,
                sample="dummy/dummy.mp3",
                cover_image="dummy/dummy.jpg",
                label=label

            )

        except IntegrityError:
            return "import failed. Your file may contain empty cells"
